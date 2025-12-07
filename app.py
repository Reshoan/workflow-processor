%%writefile app.py
from pyngrok import ngrok
import streamlit as st
import pandas as pd
from io import BytesIO
import re
import time
from openpyxl import load_workbook, Workbook
from collections import deque, defaultdict
import json
import zipfile
import os

# Try to import graphviz python package
try:
    import graphviz
    _GRAPHVIZ_AVAILABLE = True
except Exception:
    _GRAPHVIZ_AVAILABLE = False

st.set_page_config(layout="wide")
st.title("Workflow Processor and Flowchart Generator")

# ------------------------------
# Helper Functions
# ------------------------------

def write_df_to_sheet(wb, sheet_name, df, start_row=1, header=True):
    """Utility: write pandas DataFrame to openpyxl workbook sheet."""
    if sheet_name in wb.sheetnames:
        # ensure unique as safety, but normally caller manages names
        base = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base}_{counter}"
            counter += 1
    ws = wb.create_sheet(title=sheet_name)
    row0 = start_row
    if header:
        for c_idx, col in enumerate(df.columns, start=1):
            ws.cell(row=row0, column=c_idx, value=col)
        row0 += 1
    for r_idx, row in enumerate(df.values, start=row0):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    return sheet_name

def process_form_sheets(output_wb, df, field_mapping_df, selected_category_cols, created_set):
    """Create individual form sheets into output_wb. Record created sheet names."""
    from openpyxl.styles import PatternFill
    import re

    created_sheet_names = []
    all_forms = sorted(df["formName"].dropna().unique())

    # Columns from formFields to include (API-related)
    api_columns = [
        "api_source", "parameter_type", "request_parameter", "extra_request_parameter",
        "response_parameter", "action_url", "is_triggerable",
        "is_onload_api", "is_api_auto_load"
    ]

    for full_form_name in all_forms:

        # ------------------------------------------
        # Generate safe sheet name
        # ------------------------------------------
        parts = full_form_name.split("_")
        short_name = "_".join(parts[2:-1]) if len(parts) > 3 else full_form_name
        safe_name = re.sub(r'[\\/*?:\[\]]', '_', short_name)[:31]

        original_safe_name = safe_name
        counter = 1
        while safe_name in output_wb.sheetnames:
            safe_name = f"{original_safe_name}_{counter}"
            counter += 1

        # ------------------------------------------
        # Prepare data
        # ------------------------------------------
        form_df = df[df["formName"] == full_form_name].sort_values("order")

        base_columns = ["name", "input_type", "order", "data"]

        # Only keep columns that actually exist
        api_cols_existing = [c for c in api_columns if c in form_df.columns]
        base_cols_existing = [c for c in base_columns if c in form_df.columns]

        final_df = form_df[base_cols_existing + api_cols_existing].copy()

        # ------------------------------------------
        # Add Repeater column
        # ------------------------------------------
        mapped_fields = set(
            field_mapping_df[field_mapping_df['formName'] == full_form_name]['fieldName'].dropna()
        )
        final_df["Repeater"] = ["Yes" if f in mapped_fields else "" for f in final_df["name"]]

        # ------------------------------------------
        # Add Categories column
        # ------------------------------------------
        if selected_category_cols:
            categories_list = []
            for _, row in form_df.iterrows():
                selected = [c for c in selected_category_cols if row.get(c) == 1]
                categories_list.append(",".join(sorted(selected)) if selected else "")
            final_df["Categories"] = categories_list

        # ------------------------------------------
        # Reorder columns with the gap
        # ------------------------------------------
        reordered_cols = (
            ["name", "input_type", "order", "data", "Repeater", "Categories"] +
            [""] +  # empty column placeholder
            api_cols_existing
        )

        # Remove items not in final_df (safe filtering)
        reordered_cols = [c for c in reordered_cols if c == "" or c in final_df.columns]

        # Insert placeholder for empty column
        final_df[""] = ""

        final_df = final_df[reordered_cols]

        # ------------------------------------------
        # Write DataFrame starting at row 3 (gap after form title)
        # ------------------------------------------
        write_df_to_sheet(output_wb, safe_name, final_df, start_row=3, header=True)

        ws = output_wb[safe_name]

        # ------------------------------------------
        # Add form title at row 1
        # ------------------------------------------
        ws.cell(row=1, column=1, value="Form")
        ws.cell(row=1, column=2, value=full_form_name)
        # Row 2 remains empty intentionally

        # ------------------------------------------
        # Apply Column Header Styling
        # ------------------------------------------
        LIGHT_BLUE = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        LIGHT_GREEN = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        header_row = 3

        for col_idx, col_name in enumerate(reordered_cols, start=1):

            cell = ws.cell(row=header_row, column=col_idx)

            if col_name in ["name", "input_type", "order", "data", "Repeater", "Categories"]:
                cell.fill = LIGHT_BLUE

            elif col_name in api_cols_existing:
                cell.fill = LIGHT_GREEN

            # Skip the empty "" column

        # ------------------------------------------
        # Track created sheet
        # ------------------------------------------
        created_sheet_names.append(safe_name)
        created_set.add(safe_name)

    return created_sheet_names


def process_transition_and_visio(output_wb, uploaded_file, created_set):
    """Process transitions, build BFS ordering, create Transition_mapping & Transition_mapping_visio in output_wb.
       Returns node_order, edges, diagram_rows, start_node, original_name."""
    xls = pd.ExcelFile(uploaded_file)
    node_order, edges, diagram_rows, start_node = [], defaultdict(list), [], None
    original_name = uploaded_file.name

    if "transitions" not in xls.sheet_names:
        return node_order, edges, diagram_rows, start_node, original_name

    transitions_df = pd.read_excel(uploaded_file, sheet_name="transitions")

    # Ensure SLA and fromToStateName columns exist
    for col in ["sla_time", "sla_time_type", "fromToStateName"]:
        if col not in transitions_df.columns:
            transitions_df[col] = ""

    # Update 'name' based on workflowToStateName
    for idx, row in transitions_df.iterrows():
        to_state = str(row.get("workflowToStateName", "")).strip()
        current_name = str(row.get("name", "")).strip()
        if to_state == "Reject" and current_name != "Reject":
            transitions_df.at[idx, "name"] = "Reject"
        elif to_state == "Closed" and current_name != "Closed":
            transitions_df.at[idx, "name"] = "Closed"

    # Ensure 'workflowFormName' exists
    if "workflowFormName" not in transitions_df.columns:
        transitions_df["workflowFormName"] = ""

    # Desired column order with the new column
    final_order = ["name", "workflowFromStateName", "workflowToStateName",
                   "workflowFormName", "sla_time", "sla_time_type", "fromToStateName"]
    transitions_df = transitions_df[[c for c in final_order if c in transitions_df.columns]]

    # Merge SLA if available
    if "workflowSlas" in xls.sheet_names:
        slas_df = pd.read_excel(uploaded_file, sheet_name="workflowSlas")
        if {"workflowFromStateName", "workflowToStateName", "sla_time", "sla_time_type"}.issubset(slas_df.columns):
            for idx, row in transitions_df.iterrows():
                match = slas_df[
                    (slas_df["workflowFromStateName"] == row["workflowFromStateName"]) &
                    (slas_df["workflowToStateName"] == row["workflowToStateName"])
                ]
                if not match.empty:
                    transitions_df.at[idx, "sla_time"] = match.iloc[0]["sla_time"]
                    transitions_df.at[idx, "sla_time_type"] = match.iloc[0]["sla_time_type"]

        # Sort by workflowFromStateName then workflowToStateName (case-insensitive Excel-like)
        transitions_df.sort_values(
            by=["workflowFromStateName", "workflowToStateName"],
            key=lambda col: col.str.lower(),
            inplace=True
        )

    # Save Transition_mapping sheet into output_wb
    trans_sheet_name = "Transition_mapping"
    write_df_to_sheet(output_wb, trans_sheet_name, transitions_df, start_row=1, header=True)
    created_set.add(trans_sheet_name)

    # Build nodes and edges
    nodes = set()
    for _, row in transitions_df.iterrows():
        frm = str(row.get("workflowFromStateName","")).strip()
        to = str(row.get("workflowToStateName","")).strip()
        name = str(row.get("name","")).strip()
        if frm:
            nodes.add(frm)
        if to:
            nodes.add(to)
        edges[frm].append((to, name))

    # Determine start node
    if "workflowStates" in xls.sheet_names:
        workflow_df = pd.read_excel(uploaded_file, sheet_name="workflowStates")
        if "alias" in workflow_df.columns and "resolutionName" in workflow_df.columns:
            match = workflow_df[workflow_df["resolutionName"] == "Ticket Initiate"]
            if not match.empty:
                start_node = str(match.iloc[0]["alias"]).strip()
    if not start_node:
        start_node = sorted(nodes)[0] if nodes else None

    # BFS ordering
    seen = set()
    if start_node:
        q = deque([start_node])
        seen.add(start_node)
        while q:
            n = q.popleft()
            node_order.append(n)
            for to, _ in edges.get(n, []):
                if to and to not in seen:
                    seen.add(to)
                    q.append(to)
        for n in sorted(nodes):
            if n not in seen:
                node_order.append(n)
    else:
        node_order = sorted(nodes)

    # Assign IDs
    id_map = {node: f"P{(idx+1)*100}" for idx, node in enumerate(node_order)}

    # Create Visio sheet rows
    visio_rows = []
    headers = ["Process Step ID","Process Step Description","Next Step ID","Connector Label","Shape Type","Alt Text"]
    for node in node_order:
        desc = node
        targets = edges.get(node, [])
        if targets:
            tgt_map = defaultdict(list)
            for to, tname in targets:
                if to:
                    tgt_map[to].append(tname if tname else "")
            ordered_targets = sorted(tgt_map.keys(), key=lambda s: s.lower())
            next_ids = [id_map.get(t,"") for t in ordered_targets]
            connector_labels = [", ".join([n for n in tgt_map[t] if n]) for t in ordered_targets]
            next_ids_str = ",".join(next_ids) if next_ids else ""
            connector_str = ",".join(connector_labels) if connector_labels else ""
        else:
            next_ids_str = connector_str = ""

        shape = "Start" if node == start_node else "End" if not targets else "Process"
        visio_rows.append({
            "Process Step ID": id_map.get(node,""),
            "Process Step Description": desc,
            "Next Step ID": next_ids_str,
            "Connector Label": connector_str,
            "Shape Type": shape,
            "Alt Text": ""
        })

    # Write visio rows to a DataFrame then to sheet
    visio_df = pd.DataFrame(visio_rows, columns=headers)
    visio_sheet_name = "Transition_mapping_visio"
    write_df_to_sheet(output_wb, visio_sheet_name, visio_df, start_row=1, header=True)
    created_set.add(visio_sheet_name)

    # 🔹 Hide the Visio sheet
    output_wb[visio_sheet_name].sheet_state = "hidden"


    return node_order, edges, visio_rows, start_node, original_name

def generate_flowchart(diagram_rows, orientation="Vertical", file_format="PDF", workflow_name=None):
    """Generate flowchart using Graphviz with workflow title. Returns bytes."""
    if not _GRAPHVIZ_AVAILABLE:
        raise RuntimeError("graphviz package is not available in the environment.")
    dot_format = file_format.lower()
    dot = graphviz.Digraph(format=dot_format)
    dot.attr(rankdir="TB" if orientation=="Vertical" else "LR")

    # Title ABOVE the flowchart
    if workflow_name:
        dot.attr(
            label=f"Workflow: {workflow_name}",
            labelloc="t",
            labeljust="c",
            fontsize="20",
            fontname="Helvetica-Bold",
            margin="0.7"
        )
    else:
        dot.attr(margin="0.5")

    # Nodes
    for row in diagram_rows:
        node_id = row["Process Step ID"]
        label = row["Process Step Description"]
        shape = row["Shape Type"]

        fillcolor = (
            "lightyellow" if shape == "Start"
            else "lightcoral" if shape == "End"
            else "lightblue"
        )
        node_shape = (
            "oval" if shape == "Start"
            else "doublecircle" if shape == "End"
            else "box"
        )
        # Use safe node_id (Graphviz does not like empty IDs)
        safe_id = node_id if node_id else f"ID_{abs(hash(label)) % (10**8)}"
        dot.node(safe_id, label=label, shape=node_shape, style="filled", fillcolor=fillcolor)

    # Edges
    for row in diagram_rows:
        src = row["Process Step ID"]
        src_safe = src if src else f"ID_{abs(hash(row['Process Step Description'])) % (10**8)}"
        if row.get("Next Step ID"):
            targets = [t.strip() for t in str(row["Next Step ID"]).split(",") if t.strip()]
            labels = [l.strip() for l in str(row.get("Connector Label","")).split(",")]
            # ensure labels list same length as targets
            if len(labels) < len(targets):
                labels += [""] * (len(targets) - len(labels))
            for t, l in zip(targets, labels):
                tgt_safe = t if t else f"ID_{abs(hash(l)) % (10**8)}"
                dot.edge(src_safe, tgt_safe, label=l)

    # Render to temporary file
    tmp_output = "/tmp/flowchart"
    rendered_path = dot.render(filename=tmp_output, cleanup=True)

    with open(rendered_path, "rb") as fh:
        flowchart_bytes = fh.read()

    # remove the rendered file from /tmp if present
    try:
        os.remove(rendered_path)
    except Exception:
        pass

    return flowchart_bytes

def parse_from_to(value):
    """
    Parse fromToStateName formats like:
      "X"-"Y"   OR   X-Y   OR   "X" - Y
    Returns tuple (from_state, to_state) (strings) or ("","") if not parseable.
    """
    if pd.isna(value):
        return ("","")
    s = str(value).strip()
    m = re.search(r'["\']?\s*(.*?)\s*["\']?\s*-\s*["\']?\s*(.*?)\s*["\']?$', s)
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    if "-" in s:
        parts = s.split("-", 1)
        return (parts[0].strip().strip('"').strip("'"), parts[1].strip().strip('"').strip("'"))
    return ("","")

def generate_notification_summary_sheet(output_wb, uploaded_file, created_set):
    """Create notification_summary sheet in output_wb (if applicable)."""
    try:
        xls = pd.ExcelFile(uploaded_file)

        if "notifications" not in xls.sheet_names:
            return []

        notif_df = pd.read_excel(uploaded_file, sheet_name="notifications", dtype=object)

        notif_rows = []

        for _, notif_row in notif_df.iterrows():
            is_enabled = notif_row.get("is_enable_notification", None)
            if is_enabled not in [1, "1", True, "true", "True"]:
                continue

            recipients_raw = notif_row.get("recipients", None)
            if recipients_raw is None or (isinstance(recipients_raw, float) and pd.isna(recipients_raw)):
                continue

            # parse JSON safely
            try:
                if isinstance(recipients_raw, str):
                    recips = json.loads(recipients_raw)
                elif isinstance(recipients_raw, dict):
                    recips = recipients_raw
                else:
                    recips = json.loads(str(recipients_raw))
            except:
                continue

            # iterate recipients
            for receiver_name, receiver_obj in (recips.items() if isinstance(recips, dict) else []):
                if not isinstance(receiver_obj, dict):
                    continue

                rec_flag = receiver_obj.get("is_enable_notification", False)
                if rec_flag not in [True, "true", "True", 1, "1"]:
                    continue

                subject = receiver_obj.get("subject", "") or ""
                body = (
                    receiver_obj.get("email", "")
                    or receiver_obj.get("email_template", "")
                    or receiver_obj.get("body", "")
                    or ""
                )

                # parse attachments
                attachments_val = receiver_obj.get("attachments", [])
                attach_names = []

                if isinstance(attachments_val, list):
                    for a in attachments_val:
                        if isinstance(a, dict):
                            nm = a.get("name") or a.get("filename") or a.get("field") or ""
                            if nm:
                                attach_names.append(str(nm))
                        else:
                            attach_names.append(str(a))
                elif isinstance(attachments_val, dict):
                    for k, v in attachments_val.items():
                        if isinstance(v, dict) and v.get("name"):
                            attach_names.append(str(v.get("name")))
                        else:
                            attach_names.append(str(k))

                attachments_str = ",".join(attach_names)

                # parse from/to
                from_state, to_state = parse_from_to(notif_row.get("fromToStateName", ""))

                notif_rows.append({
                    "From State Name": from_state,
                    "To State Name": to_state,
                    "Receiver": receiver_name,
                    "Subject": subject,
                    "Body": body,
                    "Attachment": attachments_str
                })

        if notif_rows:
            nws_df = pd.DataFrame(notif_rows, columns=["From State Name", "To State Name", "Receiver", "Subject", "Body", "Attachment"])
            sheet_name = "notification_summary"
            write_df_to_sheet(output_wb, sheet_name, nws_df, start_row=1, header=True)
            created_set.add(sheet_name)
            return [sheet_name]

        return []
    except Exception as e:
        st.warning(f"notification_summary generation encountered an issue: {e}")
        return []

def column_remover(original_wb, output_wb, created_set):
    """
    For each configured sheet, create a modified version (only keep specific columns)
    in output_wb and mark them as created (because they are 'modified' versions).
    """
    sheet_configs = {
        "workflowStates": ["name", "alias", "actor", "access_type"],
        "transitionAction": ["action_method", "parameters", "transitionName"],
        "autoTransition": ["is_enable", "transitionName", "workflowFromStateName", "workflowToStateName"],
        "downloadExcelConfigs": ["report_id", "btn_title", "workflowStateName"],
        "workflowStatesGroup": ["workflow_id", "state_id", "group_id", "create", "update", "view", "delete", "action", "draft"]
    }

    for sheet_name, keep_cols in sheet_configs.items():
        if sheet_name not in original_wb.sheetnames:
            continue

        # Read sheet into DataFrame
        ws = original_wb[sheet_name]
        # Extract header values
        headers = [cell.value for cell in ws[1] if cell.value]
        if not headers:
            continue

        # Build dataframe rows
        rows = []
        for r in range(2, ws.max_row + 1):
            row_vals = []
            for h in headers:
                val = ws.cell(row=r, column=headers.index(h) + 1).value
                row_vals.append(val)
            rows.append(row_vals)

        if rows:
            df = pd.DataFrame(rows, columns=headers)
        else:
            df = pd.DataFrame(columns=headers)

        # Keep only the configured columns that exist in the original
        keep_existing = [c for c in keep_cols if c in df.columns]
        if not keep_existing:
            continue

        filtered_df = df[keep_existing].copy()

        # Write filtered_df to output_wb as the modified sheet (same sheet name)
        final_name = sheet_name
        if final_name in output_wb.sheetnames:
            base = final_name
            counter = 1
            while final_name in output_wb.sheetnames:
                final_name = f"{base}_{counter}"
                counter += 1

        write_df_to_sheet(output_wb, final_name, filtered_df, start_row=1, header=True)
        created_set.add(final_name)

def extract_workflow_name_from_file(uploaded_file):
    """
    Reads the 'workflow' sheet -> 'name' column -> removes everything after last underscore.
    Fallback: return filename (without extension).
    """

    try:
        uploaded_file.seek(0)
        xls = pd.ExcelFile(uploaded_file)

        if "workflow" not in xls.sheet_names:
            return os.path.splitext(uploaded_file.name)[0]

        uploaded_file.seek(0)
        wf_df = pd.read_excel(uploaded_file, sheet_name="workflow")

        if "name" not in wf_df.columns or wf_df["name"].dropna().empty:
            return os.path.splitext(uploaded_file.name)[0]

        raw_value = str(wf_df["name"].dropna().iloc[0]).strip()

        # Remove everything after final underscore
        if "_" in raw_value:
            return "_".join(raw_value.split("_")[:-1])

        return raw_value

    except Exception:
        # If anything fails, fallback to filename without extension
        return os.path.splitext(uploaded_file.name)[0]

def enrich_workflow_group(output_wb, dwe_group_df):
    """
    After column_remover has created workflowStatesGroup,
    use dwe_group_list to enrich it with name + admin_user_ids.
    """

    if "workflowStatesGroup" not in output_wb.sheetnames:
        return  # Nothing to enrich

    ws = output_wb["workflowStatesGroup"]

    # Extract headers
    headers = [cell.value for cell in ws[1]]

    # Ensure workflow_id exists
    if "workflow_id" not in headers:
        return

    workflow_id_col = headers.index("workflow_id") + 1
    draft_col_idx = headers.index("draft") + 1

    # Get first workflow_id
    workflow_id = ws.cell(row=2, column=workflow_id_col).value
    if workflow_id is None:
        return

    # Search for this ID inside dwe_group_list
    match_row = dwe_group_df.loc[dwe_group_df["# id"] == workflow_id]

    if match_row.empty:
        return

    name_value = match_row["name"].iloc[0]
    user_value = match_row["admin_user_ids"].iloc[0]

    # Insert a blank column after draft
    ws.insert_cols(draft_col_idx + 1)

    # Insert name + users columns beside it
    ws.insert_cols(draft_col_idx + 2)
    ws.insert_cols(draft_col_idx + 3)

    ws.cell(row=1, column=draft_col_idx + 2).value = "name"
    ws.cell(row=1, column=draft_col_idx + 3).value = "users"

    # Only fill in ONE row (row 2)
    target_row = 2
    ws.cell(row=target_row, column=draft_col_idx + 2).value = name_value
    ws.cell(row=target_row, column=draft_col_idx + 3).value = user_value


# ------------------------------
# Main Execution (Streamlit UI) - BULK UPLOAD WITH START BUTTON (clean uploader)
# ------------------------------

# Initialize session state keys
for key in [
    "uploaded_files",            # list of uploaded files
    "master_zip_bytes",          # final zip containing all workflows
    "bulk_processing_errors",    # errors per-file
    "bulk_created_workflows"     # list of processed workflow names
]:
    if key not in st.session_state:
        st.session_state[key] = []

# Step 1: Upload files (users can remove files via uploader's built-in 'x')
uploaded_files = st.file_uploader(
    "Upload one or more Excel workflow files",
    type=["xlsx"],
    accept_multiple_files=True,
    key="file_uploader"
)

# Sync session state with current uploader selection
if uploaded_files:
    st.session_state.uploaded_files = uploaded_files
else:
    st.session_state.uploaded_files = []


# Step 2: Show number of uploaded files
if st.session_state.uploaded_files:
    st.subheader(f"{len(st.session_state.uploaded_files)} file(s) ready for processing")

    # Step 3: Start processing button
    if st.button("Start Processing"):
        if not st.session_state.uploaded_files:
            st.warning("No files selected for processing.")
        else:
            # Reset previous results
            st.session_state.master_zip_bytes = None
            st.session_state.bulk_processing_errors = None
            st.session_state.bulk_created_workflows = None

            # ------------------------------
            # Begin bulk processing
            # ------------------------------
            files_to_process = st.session_state.uploaded_files
            st.info(f"{len(files_to_process)} file(s) selected — starting bulk processing.")
            progress_outer = st.progress(0)
            status_box = st.empty()
            errors = {}
            processed_workflows = []

            # Track duplicate workflow names
            workflow_name_counter = {}


            # Detect if dwe_group_list is uploaded (used later for enrichment)
            dwe_group_df = None
            for f in files_to_process:
                if f.name.lower().startswith("dwe_group_list"):
                    f.seek(0)
                    dwe_group_df = pd.read_excel(f, sheet_name="Sheet1")
                    break
            # Notify user whether dwe_group_list.xlsx was found
            if dwe_group_df is not None:
                st.success("✓ 'dwe_group_list.xlsx' detected — enrichment will be applied.")
            else:
                st.warning("⚠ No 'dwe_group_list.xlsx' found — workflowStatesGroup enrichment will be skipped.")



            master_zip_buf = BytesIO()
            with zipfile.ZipFile(master_zip_buf, "w", zipfile.ZIP_DEFLATED) as master_zip:

                for idx, uploaded_file in enumerate(files_to_process):
                    try:
                        status_box.info(f"Processing file {idx+1}/{len(files_to_process)}: {uploaded_file.name}")
                        # Skip dwe_group_list entirely
                        if uploaded_file.name.lower().startswith("dwe_group_list"):
                            continue

                        progress_outer.progress(int((idx / len(files_to_process)) * 100))

                        uploaded_file.seek(0)
                        workflow_name_trimmed = extract_workflow_name_from_file(uploaded_file)

                        # Sanitize base name
                        base_safe_name = re.sub(r'[\\/*?:\[\]]', '_', workflow_name_trimmed) or f"workflow_{idx+1}"

                        # Handle duplicate workflow names
                        if base_safe_name in workflow_name_counter:
                            workflow_name_counter[base_safe_name] += 1
                            safe_workflow_folder = f"{base_safe_name}_{workflow_name_counter[base_safe_name]}"
                        else:
                            workflow_name_counter[base_safe_name] = 0
                            safe_workflow_folder = base_safe_name


                        # --- All existing processing logic remains unchanged ---
                        output_wb = Workbook()
                        if output_wb.active and output_wb.active.title == "Sheet":
                            output_wb.remove(output_wb.active)

                        created_or_modified = set()
                        created_sheet_names = []

                        uploaded_file.seek(0)
                        xls = pd.ExcelFile(uploaded_file)
                        if "formFields" not in xls.sheet_names:
                            note = f"The file '{uploaded_file.name}' does not contain a 'formFields' sheet. Skipping detailed processing."
                            master_zip.writestr(f"{safe_workflow_folder}/ERROR.txt", note)
                            errors[uploaded_file.name] = "Missing 'formFields' sheet"
                            continue

                        uploaded_file.seek(0)
                        df = pd.read_excel(uploaded_file, sheet_name="formFields")
                        selected_category_cols = [c for c in [
                            "is_editable","is_hidden","is_required","is_nullable",
                            "is_multiselect","is_richtext","editor_height",
                            "is_encrypted","is_conditional"
                        ] if c in df.columns]

                        field_mapping_df = pd.DataFrame(columns=["formName","fieldName"])
                        if "fieldMapping" in xls.sheet_names:
                            uploaded_file.seek(0)
                            field_mapping_df = pd.read_excel(uploaded_file, sheet_name="fieldMapping")

                        required_columns = {"formName", "fieldName"}
                        if "fieldMapping" in xls.sheet_names:
                            missing_cols = required_columns - set(field_mapping_df.columns)
                            if missing_cols:
                                raise RuntimeError(
                                    f"The 'fieldMapping' sheet is missing required columns: {', '.join(missing_cols)}."
                                )

                        created_by_forms = process_form_sheets(output_wb, df, field_mapping_df, selected_category_cols, created_or_modified)
                        created_sheet_names += created_by_forms

                        uploaded_file.seek(0)
                        node_order, edges, diagram_rows, start_node, original_name = process_transition_and_visio(output_wb, uploaded_file, created_or_modified)
                        if "Transition_mapping" in created_or_modified:
                            created_sheet_names.append("Transition_mapping")
                        if "Transition_mapping_visio" in created_or_modified:
                            created_sheet_names.append("Transition_mapping_visio")

                        uploaded_file.seek(0)
                        notif_created = generate_notification_summary_sheet(output_wb, uploaded_file, created_or_modified)
                        created_sheet_names += notif_created

                        uploaded_file.seek(0)
                        original_wb = load_workbook(uploaded_file, data_only=True)
                        column_remover(original_wb, output_wb, created_or_modified)
                        created_sheet_names += sorted([n for n in created_or_modified if n not in created_sheet_names])
                        # Add name + users enrichment to workflowStatesGroup
                        if dwe_group_df is not None:
                            enrich_workflow_group(output_wb, dwe_group_df)


                        # --- Flowchart embed & save workbook logic remains unchanged ---
                        try:
                            from openpyxl.drawing.image import Image as XLImage
                            from PIL import Image as PILImage
                            import io as _io

                            if diagram_rows:
                                fc_png = generate_flowchart(
                                    diagram_rows,
                                    orientation="Horizontal",
                                    file_format="PNG",
                                    workflow_name=workflow_name_trimmed
                                )

                                temp_png_path = f"/tmp/flowchart_embed_{idx}.png"
                                pil_img = PILImage.open(_io.BytesIO(fc_png))
                                pil_img.save(temp_png_path)

                                if "flowchart" in output_wb.sheetnames:
                                    del output_wb["flowchart"]
                                ws_flow = output_wb.create_sheet("flowchart")
                                excel_img = XLImage(temp_png_path)
                                ws_flow.add_image(excel_img, "A1")

                                created_or_modified.add("flowchart")
                                created_sheet_names.append("flowchart")

                        except Exception as e:
                            st.warning(f"[{uploaded_file.name}] Failed to embed horizontal flowchart into Excel: {e}")

                        processed_bytes_io = BytesIO()
                        output_wb.save(processed_bytes_io)
                        processed_bytes_io.seek(0)
                        processed_bytes = processed_bytes_io.getvalue()

                        original_excel_name = f"{workflow_name_trimmed}_Original.xlsx"
                        processed_excel_name = f"Processed_{workflow_name_trimmed}.xlsx"
                        folder_prefix = f"{safe_workflow_folder}/"

                        uploaded_file.seek(0)
                        original_data = uploaded_file.getvalue()
                        master_zip.writestr(folder_prefix + original_excel_name, original_data)
                        master_zip.writestr(folder_prefix + processed_excel_name, processed_bytes)

                        if diagram_rows:
                            for fmt in ["PDF", "SVG", "PNG"]:
                                for orient in ["Vertical", "Horizontal"]:
                                    try:
                                        fc_bytes = generate_flowchart(
                                            diagram_rows,
                                            orientation=orient,
                                            file_format=fmt,
                                            workflow_name=workflow_name_trimmed
                                        )
                                        orient_flag = "V" if orient == "Vertical" else "H"
                                        ext = fmt.lower()
                                        fc_name = f"{workflow_name_trimmed}_Flowchart_{orient_flag}.{ext}"
                                        master_zip.writestr(folder_prefix + fc_name, fc_bytes)
                                    except Exception as e:
                                        note = f"Flowchart generation failed for {fmt} {orient}: {e}"
                                        master_zip.writestr(folder_prefix + f"FLOWCHART_ERROR_{fmt}_{orient}.txt", note)
                        else:
                            master_zip.writestr(folder_prefix + "NO_FLOWCHARTS.txt", "No transition/diagram rows found to generate flowcharts.")

                        readme_lines = [
                            f"Workflow: {workflow_name_trimmed}",
                            f"Original filename: {uploaded_file.name}",
                            "",
                            "Files generated:",
                            f"- {original_excel_name}",
                            f"- {processed_excel_name}",
                        ]
                        if diagram_rows:
                            readme_lines.append("- Flowcharts (PDF, SVG, PNG)")
                        else:
                            readme_lines.append("- No flowcharts generated.")

                        readme_lines.append("")
                        readme_lines.append("Created/modified sheets:")
                        for i, s in enumerate(sorted(created_or_modified), start=1):
                            readme_lines.append(f"{i}. {s}")

                        master_zip.writestr(folder_prefix + "README.txt", "\n".join(readme_lines))
                        processed_workflows.append(workflow_name_trimmed)

                    except Exception as e:
                        err_msg = f"Error processing '{uploaded_file.name}': {e}"
                        errors[uploaded_file.name] = str(e)
                        master_zip.writestr(f"ERRORS_{uploaded_file.name}.txt", err_msg)

                progress_outer.progress(100)

            master_zip_buf.seek(0)
            st.session_state.master_zip_bytes = master_zip_buf.getvalue()
            st.session_state.bulk_processing_errors = errors
            st.session_state.bulk_created_workflows = processed_workflows

            if processed_workflows:
                st.success(f"Processed {len(processed_workflows)} workflow(s): {', '.join(processed_workflows)}")
            if errors:
                st.error(f"Errors occurred for {len(errors)} file(s). See ZIP for details.")

# ------------------------------
# Download area
# ------------------------------
if st.session_state.master_zip_bytes:
    st.subheader("Download bulk result")
    st.download_button(
        label="Download all workflows (ZIP)",
        data=st.session_state.master_zip_bytes,
        file_name="workflows_bundle.zip",
        mime="application/zip"
    )

# ------------------------------
# End of file
# ------------------------------
